from openpyxl import load_workbook

# Load the workbook
wb = load_workbook("C:\PythonWorkSpace\PythonBasicsLearnings\ExcelOperations\Files\Studentdata.xlsx")

# Select a sheet
sheet = wb["Sheet1"]

# Read specific cell
print("A1 value -->>"+sheet["A1"].value)

# Read all rows
for row in sheet.iter_rows(values_only=True):
    print(row)